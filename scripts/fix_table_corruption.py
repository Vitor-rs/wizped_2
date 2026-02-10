"""
Fix corrupted Excel table definition in ficha_freq_gerador.xlsm.
The backfill_professores.py renamed the header of BD_Alunos col 9 from
'ID_Professor' to 'Professores', which broke table8.xml because openpyxl
didn't update the table column definition.

This script:
1. Opens the workbook
2. Finds the table on BD_Alunos
3. Fixes the table column name to match the actual header
4. Ensures the table range covers all data rows
5. Saves the file
"""

import openpyxl
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter

WORKBOOK_PATH = r"c:\Users\user\Documents\GitHub\wizped_2\ficha_freq_gerador.xlsm"


def main():
    print(f"Opening {WORKBOOK_PATH}...")
    wb = openpyxl.load_workbook(WORKBOOK_PATH, keep_vba=True)

    # List all tables in all sheets
    print("\n=== ALL TABLES ===")
    table_count = 0
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        if hasattr(ws, "tables") and ws.tables:
            for tbl_name in ws.tables:
                tbl = ws.tables[tbl_name]
                table_count += 1
                print(
                    f"  Table #{table_count}: '{tbl_name}' on sheet '{sheet_name}' -> ref={tbl.ref}"
                )
                # Show column names
                if tbl.tableColumns:
                    col_names = [tc.name for tc in tbl.tableColumns]
                    print(f"    Columns: {col_names}")
                # Show actual headers on the sheet
                from openpyxl.utils.cell import range_boundaries

                min_col, min_row, max_col, max_row = range_boundaries(tbl.ref)
                actual_headers = []
                for c in range(min_col, max_col + 1):
                    actual_headers.append(ws.cell(min_row, c).value)
                print(f"    Actual headers: {actual_headers}")

                # Check for mismatch
                if tbl.tableColumns:
                    for i, tc in enumerate(tbl.tableColumns):
                        actual = actual_headers[i] if i < len(actual_headers) else None
                        if tc.name != actual and actual is not None:
                            print(
                                f"    *** MISMATCH col {i}: table says '{tc.name}', sheet says '{actual}'"
                            )
                            tc.name = actual
                            print(f"    *** FIXED: renamed to '{actual}'")

                # Check if table range covers all data
                actual_last_row = ws.max_row
                if max_row < actual_last_row:
                    new_ref = f"{get_column_letter(min_col)}{min_row}:{get_column_letter(max_col)}{actual_last_row}"
                    print(
                        f"    *** RANGE MISMATCH: table ends at row {max_row}, data goes to {actual_last_row}"
                    )
                    tbl.ref = new_ref
                    print(f"    *** FIXED: new ref = {new_ref}")

    if table_count == 0:
        print("  No tables found!")

    # Save
    print(f"\nSaving {WORKBOOK_PATH}...")
    wb.save(WORKBOOK_PATH)
    wb.close()
    print("Done!")


if __name__ == "__main__":
    main()
