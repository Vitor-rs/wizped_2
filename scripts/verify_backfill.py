"""Quick verify: show BD_Alunos col 9 after backfill."""

import openpyxl

wb = openpyxl.load_workbook(
    r"c:\Users\user\Documents\GitHub\wizped_2\ficha_freq_gerador.xlsm", data_only=True
)
ws = wb["BD_Alunos"]
print(f"Header col 9: {ws.cell(1, 9).value}")
print("\nSample (first 10 students with professors):")
count = 0
for r in range(2, ws.max_row + 1):
    v = ws.cell(r, 9).value
    if v is not None and v != "":
        name = ws.cell(r, 2).value
        print(f"  {name}: [{v}]")
        count += 1
        if count >= 10:
            break
print(f"\nTotal students with professors in col 9: ", end="")
total = sum(
    1 for r in range(2, ws.max_row + 1) if ws.cell(r, 9).value not in (None, "")
)
print(total)
wb.close()
