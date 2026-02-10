"""Analyze BD_Alunos, BD_Vinculo_Professor, BD_Professores from the Excel workbook."""

import openpyxl

wb = openpyxl.load_workbook(
    r"c:\Users\user\Documents\GitHub\wizped_2\ficha_freq_gerador.xlsm", data_only=True
)

print("=" * 60)
print("SHEETS IN WORKBOOK:")
print("=" * 60)
for name in wb.sheetnames:
    print(f"  - {name}")


def print_sheet(sheet_name, max_data_rows=10):
    print(f"\n{'=' * 60}")
    print(sheet_name)
    print("=" * 60)
    if sheet_name not in wb.sheetnames:
        print("  NOT FOUND!")
        return
    ws = wb[sheet_name]
    # Headers
    headers = []
    for c in range(1, ws.max_column + 1):
        val = ws.cell(1, c).value
        if val is not None:
            headers.append((c, val))
    print(f"Columns: {headers}")
    print(f"max_row: {ws.max_row}, max_column: {ws.max_column}")
    # Data rows
    count = 0
    for r in range(2, ws.max_row + 1):
        row_data = {}
        has_data = False
        for c, hdr in headers:
            v = ws.cell(r, c).value
            row_data[hdr] = v
            if v is not None:
                has_data = True
        if has_data:
            count += 1
            if count <= max_data_rows:
                print(f"  Row {r}: {row_data}")
    print(f"  Total data rows: {count}")


# Analyze key sheets
print_sheet("BD_Alunos", max_data_rows=5)
print_sheet("BD_Professores", max_data_rows=20)
print_sheet("BD_Vinculo_Professor", max_data_rows=20)
print_sheet("BD_TipoOcorrencia", max_data_rows=20)

wb.close()
print("\nDone.")
