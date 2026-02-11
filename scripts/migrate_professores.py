"""
Migração: BD_Professores → BD_Funcionarios  +  BD_Historico col 7.

Two-phase approach to avoid openpyxl table corruption:
  Phase 1 (openpyxl): Cell data only — headers, Funcao values
  Phase 2 (zipfile):  Patch table XML directly in the saved file

Uso:
    python scripts/migrate_professores.py [caminho_xlsx]
"""

import sys
import os
import re
import shutil
import zipfile
import tempfile

import openpyxl
from openpyxl.utils import get_column_letter

DEFAULT_PATH = os.path.join("docs", "ficha_freq_gerador.xlsm")
path = sys.argv[1] if len(sys.argv) > 1 else DEFAULT_PATH
path = os.path.abspath(path)
print(f"Arquivo: {path}\n")

# Backup
backup_path = path + ".bak"
shutil.copy2(path, backup_path)
print(f"Backup: {backup_path}\n")


# ═══════════════════════════════════════════════════════
# PHASE 1: Cell data via openpyxl (NO table manipulation)
# ═══════════════════════════════════════════════════════
print("=== FASE 1: Dados de célula via openpyxl ===\n")

wb = openpyxl.load_workbook(path, keep_vba=True)

# --- BD_Professores → BD_Funcionarios (sheet rename) ---
ws = None
if "BD_Professores" in wb.sheetnames:
    ws = wb["BD_Professores"]
    ws.title = "BD_Funcionarios"
    print("  ✓ Sheet renomeada: BD_Professores → BD_Funcionarios")
elif "BD_Funcionarios" in wb.sheetnames:
    ws = wb["BD_Funcionarios"]
    print("  ✓ Sheet já se chama BD_Funcionarios")
else:
    print("  ✗ Sheet BD_Professores não encontrada!")
    sys.exit(1)

# Rename header cells
if ws.cell(1, 1).value == "ID_Professor":
    ws.cell(1, 1).value = "ID_Funcionario"
    print("  ✓ Header A1: ID_Professor → ID_Funcionario")

# Add Funcao column data (col 3)
col_funcao = 3
ws.cell(1, col_funcao).value = "Funcao"
filled = 0
for r in range(2, ws.max_row + 1):
    if ws.cell(r, 1).value is not None and ws.cell(r, col_funcao).value is None:
        ws.cell(r, col_funcao).value = "Professor"
        filled += 1
print(f"  ✓ Funcao: {filled} registros preenchidos como 'Professor'")

# --- BD_Historico: add header in col 7 ---
ws_hist = wb["BD_Historico"]
ws_hist.cell(1, 7).value = "ID_Funcionario"
print("  ✓ BD_Historico: header col 7 = ID_Funcionario")

# --- BD_Vinculo_Professor: rename header ---
ws_vp = wb["BD_Vinculo_Professor"]
if ws_vp.cell(1, 3).value == "ID_Professor":
    ws_vp.cell(1, 3).value = "ID_Funcionario"
    print("  ✓ BD_Vinculo_Professor: header col 3 = ID_Funcionario")

# IMPORTANT: Do NOT touch any table objects via openpyxl!
# Save cell data only
wb.save(path)
print("  ✓ Dados de célula salvos.\n")


# ═══════════════════════════════════════════════════════
# PHASE 2: Patch table XML directly in the ZIP
# ═══════════════════════════════════════════════════════
print("=== FASE 2: Patch XML das tabelas no ZIP ===\n")


def patch_xml_text(text, replacements):
    """Apply text replacements, returns (new_text, count)."""
    total = 0
    for old, new in replacements:
        text, n = re.subn(re.escape(old), new, text)
        total += n
    return text, total


def expand_table_xml(text, new_col_name, new_col_id):
    """Add a column to table XML: expand ref, autoFilter, tableColumns, filterColumn."""

    # 1. Expand ref="A1:Xn" -> "A1:Yn"
    def bump_ref(m):
        prefix = m.group(1)
        col_letter = m.group(2)
        row = m.group(3)
        col_num = 0
        for ch in col_letter:
            col_num = col_num * 26 + (ord(ch) - ord("A") + 1)
        new_letter = get_column_letter(col_num + 1)
        return f"{prefix}{new_letter}{row}"

    text = re.sub(r'(ref="A1:)([A-Z]+)(\d+")', bump_ref, text)

    # 2. Bump tableColumns count
    m = re.search(r'tableColumns count="(\d+)"', text)
    if m:
        old_cnt = int(m.group(1))
        text = text.replace(
            f'tableColumns count="{old_cnt}"', f'tableColumns count="{old_cnt + 1}"'
        )

        # 3. Add filterColumn for new colId (match openpyxl format)
        new_filter = (
            f'<filterColumn colId="{old_cnt}" hiddenButton="1" showButton="1" />'
        )
        text = text.replace("</autoFilter>", new_filter + "</autoFilter>")

        # 4. Add tableColumn entry (no xr3:uid — openpyxl strips that namespace)
        new_col = f'<tableColumn id="{new_col_id}" name="{new_col_name}" />'
        text = text.replace("</tableColumns>", new_col + "</tableColumns>")

    return text


# Read the saved file and patch table XMLs
xml_patches = {}

with zipfile.ZipFile(path, "r") as z:
    # --- table5: Tbl_Professores → Tbl_Funcionarios + add Funcao col ---
    t5 = z.read("xl/tables/table5.xml").decode("utf-8")
    t5, n = patch_xml_text(
        t5,
        [
            ('name="Tbl_Professores"', 'name="Tbl_Funcionarios"'),
            ('displayName="Tbl_Professores"', 'displayName="Tbl_Funcionarios"'),
            ('name="ID_Professor"', 'name="ID_Funcionario"'),
        ],
    )
    t5 = expand_table_xml(t5, "Funcao", 3)
    xml_patches["xl/tables/table5.xml"] = t5.encode("utf-8")
    print("  ✓ table5: Tbl_Professores → Tbl_Funcionarios + col Funcao")

    # --- table11: Tbl_Vinculo_Professor (rename FK only) ---
    t11 = z.read("xl/tables/table11.xml").decode("utf-8")
    t11, n = patch_xml_text(
        t11,
        [
            ('name="ID_Professor"', 'name="ID_Funcionario"'),
        ],
    )
    xml_patches["xl/tables/table11.xml"] = t11.encode("utf-8")
    print("  ✓ table11: Tbl_Vinculo_Professor FK renamed")

    # --- table13: Tbl_Historico (add ID_Funcionario col) ---
    t13 = z.read("xl/tables/table13.xml").decode("utf-8")
    t13 = expand_table_xml(t13, "ID_Funcionario", 7)
    xml_patches["xl/tables/table13.xml"] = t13.encode("utf-8")
    print("  ✓ table13: Tbl_Historico + col ID_Funcionario")

# Write patched ZIP
print("\n  Escrevendo ZIP final...")
tmp_fd, tmp_path = tempfile.mkstemp(suffix=".xlsm")
os.close(tmp_fd)

with zipfile.ZipFile(path, "r") as zin:
    with zipfile.ZipFile(tmp_path, "w", compression=zipfile.ZIP_DEFLATED) as zout:
        for item in zin.infolist():
            if item.filename in xml_patches:
                zout.writestr(item, xml_patches[item.filename])
            else:
                zout.writestr(item, zin.read(item.filename))

shutil.move(tmp_path, path)
print("  ✓ ZIP patcheado salvo.\n")

print("=" * 50)
print("✓ Migração completa!")
print("Próximo passo: rodar rebuild_model.py para atualizar Data Model.")
