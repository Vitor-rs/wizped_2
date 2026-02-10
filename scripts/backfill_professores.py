"""
Backfill BD_Alunos col 9 with concatenated professor names.
Renames header from 'ID_Professor' to 'Professores'.
Uses BD_Vinculo_Professor + BD_Professores to resolve names.

Usage: python scripts/backfill_professores.py
"""

import openpyxl
import sys

WORKBOOK_PATH = r"c:\Users\user\Documents\GitHub\wizped_2\ficha_freq_gerador.xlsm"


def main():
    print(f"Opening {WORKBOOK_PATH}...")
    wb = openpyxl.load_workbook(WORKBOOK_PATH, keep_vba=True)

    ws_alunos = wb["BD_Alunos"]
    ws_vinculo = wb["BD_Vinculo_Professor"]
    ws_profs = wb["BD_Professores"]

    # 1. Build professor name lookup: {ID_Professor: Nome}
    prof_names = {}
    for r in range(2, ws_profs.max_row + 1):
        pid = ws_profs.cell(r, 1).value
        pname = ws_profs.cell(r, 2).value
        if pid is not None and pname is not None:
            prof_names[int(pid)] = str(pname)
    print(f"  Professors: {prof_names}")

    # 2. Build vinculo lookup: {ID_Aluno: [ID_Professor, ...]}
    vinculos = {}
    for r in range(2, ws_vinculo.max_row + 1):
        aluno_id = ws_vinculo.cell(r, 2).value
        prof_id = ws_vinculo.cell(r, 3).value
        if aluno_id is not None and prof_id is not None:
            aid = int(aluno_id)
            pid = int(prof_id)
            if aid not in vinculos:
                vinculos[aid] = []
            vinculos[aid].append(pid)
    print(f"  Vinculos: {len(vinculos)} students have professor bindings")

    # 3. Rename header
    old_header = ws_alunos.cell(1, 9).value
    ws_alunos.cell(1, 9).value = "Professores"
    print(f"  Header col 9: '{old_header}' -> 'Professores'")

    # 4. Backfill each student
    updated = 0
    for r in range(2, ws_alunos.max_row + 1):
        aluno_id = ws_alunos.cell(r, 1).value
        if aluno_id is None:
            continue
        aid = int(aluno_id)
        if aid in vinculos:
            names = []
            for pid in vinculos[aid]:
                if pid in prof_names:
                    names.append(prof_names[pid])
                else:
                    names.append(f"ID:{pid}")
            ws_alunos.cell(r, 9).value = ", ".join(names)
            updated += 1
        else:
            # No vinculo — check if there was an old ID_Professor value
            old_val = ws_alunos.cell(r, 9).value
            if old_val is not None and old_val != "":
                try:
                    old_pid = int(old_val)
                    if old_pid in prof_names:
                        ws_alunos.cell(r, 9).value = prof_names[old_pid]
                        updated += 1
                    else:
                        ws_alunos.cell(r, 9).value = ""
                except (ValueError, TypeError):
                    pass  # Already a string, leave as-is

    print(f"  Updated {updated} student rows")

    # 5. Save
    wb.save(WORKBOOK_PATH)
    wb.close()
    print(f"\nSaved to {WORKBOOK_PATH}")
    print("Done!")


if __name__ == "__main__":
    main()
