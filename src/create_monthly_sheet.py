import openpyxl
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter
import os

FILE_NAME = "ficha_freq_gerador.xlsx"
SHEET_NAME = "BP_FichaMensal"
TABLE_NAME = "Tbl_FichaMensal"


def create_monthly_sheet():
    file_path = os.path.abspath(FILE_NAME)
    if not os.path.exists(file_path):
        print(f"File not found: {file_path}")
        return

    print(f"Opening {FILE_NAME}...")
    try:
        wb = openpyxl.load_workbook(file_path)

        # Check if sheet exists, if so delete/recreate or just get it
        if SHEET_NAME in wb.sheetnames:
            print(
                f"Sheet {SHEET_NAME} already exists. Deleting to recreate structure..."
            )
            del wb[SHEET_NAME]

        ws = wb.create_sheet(SHEET_NAME)
        print(f"Created sheet {SHEET_NAME}")

        # Define Columns
        # 1. ID Aluno
        # 2. Nome do Aluno
        # 3. Livro
        # 4. Agenda (Dias/Horário)
        # 5. Tipo (Exp + Mod + VIP)
        # 6-36. Days 1-31

        fixed_headers = ["ID_Aluno", "Nome_Aluno", "Livro", "Agenda", "Tipo"]
        day_headers = [str(i) for i in range(1, 32)]

        headers = fixed_headers + day_headers

        # Write Headers
        ws.append(headers)

        # Create Table
        # Range: A1 : [LastCol]2 (Table needs at least one data row usually, or just headers)
        # If openpyxl allows headers only, distinct checks needed.
        # Often safest to add one empty row if strictly creating a table.
        # But let's try just headers.

        last_col_letter = get_column_letter(len(headers))
        ref = f"A1:{last_col_letter}2"  # Assuming we add an empty row for stability

        # Add empty row so table is valid
        ws.append([""] * len(headers))

        print(f"Creating table {TABLE_NAME} at {ref}")

        tab = Table(displayName=TABLE_NAME, ref=ref)

        style = TableStyleInfo(
            name="TableStyleMedium9",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False,
        )
        tab.tableStyleInfo = style

        ws.add_table(tab)

        # Adjust column widths (optional logic)
        ws.column_dimensions["B"].width = 30  # Nome
        ws.column_dimensions["C"].width = 20  # Livro
        ws.column_dimensions["D"].width = 25  # Agenda

        # Days narrow
        for i in range(6, 37):  # Columns F onwards
            col_letter = get_column_letter(i)
            ws.column_dimensions[col_letter].width = 4

        wb.save(file_path)
        print("Success! Annual sheet structure created.")

    except Exception as e:
        print(f"Error: {e}")


if __name__ == "__main__":
    create_monthly_sheet()
